#!/usr/bin/env python
# coding: utf-8

# In[114]:


#pip install openpyxl==3.1.3


# In[115]:


import pandas as pd
import csv
import openpyxl
from openpyxl.chart import DoughnutChart, Reference
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.fill import PatternFillProperties
from openpyxl.drawing.line import LineProperties
import plotly.express as px
import plotly
from openpyxl.chart import LineChart, Reference


# In[116]:


tdb_obs = pd.DataFrame()


# In[117]:


with pd.ExcelWriter('data/tdb_obs.xlsx', engine='openpyxl') as writer:
    tdb_obs.to_excel(writer, sheet_name='Feuil1', index=False)


# # Synonymes

# In[118]:


"""Synonymes"""

df_synonymes = pd.read_csv('data/synonymes.csv')

# # L’écrire dans un fichier Excel
# df_synonymes.to_excel('data/synonymes.xlsx', index=False)


# In[119]:


# Trier sur la colonne 'proximite_mot' et prendre les 7 plus grands
df_top7 = df_synonymes.sort_values(by='score_proximite_mot', ascending=False).head(7)
df_top7


# In[120]:


# Réécriture de la feuille Excel dans une feuille temporaire de travail
with pd.ExcelWriter('data/tdb_obs.xlsx', mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    df_top7.to_excel(writer, sheet_name='synonymes_top7', index=False)


# In[121]:


# Charger le fichier Excel
wb = openpyxl.load_workbook('data/tdb_obs.xlsx')

# Aller sur la feuille principale
main_sheet = wb['Feuil1']

# Aller chercher les données triées dans la feuille 'synonymes_top7'
syn_sheet = wb['synonymes_top7']

# Création du graphique
chart = BarChart()
chart.type = "bar"
chart.style = 11
chart.title = "Synonymes"
chart.y_axis.title = ' '
chart.x_axis.title = ' '

# Les catégories (noms des synonymes en B)
cats = Reference(syn_sheet, min_col=2, min_row=2, max_row=8)

# Les données (scores en C)
data = Reference(syn_sheet, min_col=3, min_row=1, max_row=8)

chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

chart.legent = None

# On insère le graphique dans la feuille principale (Feuil1)
main_sheet.add_chart(chart, "A10")

# Sauvegarde
wb.save('data/tdb_obs.xlsx')
wb.close()


# 

# In[122]:


"""# Pie charts sur le mot dans les arts"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference

df_mot_in_arts = pd.read_csv("data/repartition_types_oeuvres_art.csv")

# Créer une feuille pour les données
ws_data = wb.create_sheet(title="repartition_titre_type_oeuvre")

# Écrire les données du DataFrame dans la feuille de données
ws_data.append(['type', 'nombre'])
for _, row in df_mot_in_arts.iterrows():
    ws_data.append([row['type'], row['nombre']])

# Créer une feuille pour le graphique
ws_chart = wb["Feuil1"]

# Créer un graphique en secteurs
pie = PieChart()

# Définir les références aux données et catégories
data_ref = Reference(ws_data, min_col=2, min_row=2, max_row=4, max_col=2)
categories_ref = Reference(ws_data, min_col=1, min_row=2, max_row=4)

# Ajouter les données et les catégories au graphique
pie.add_data(data_ref)
pie.set_categories(categories_ref)

# Ajouter le graphique à la feuille de graphique
ws_chart.add_chart(pie, "H2")

wb.save('data/tdb_obs.xlsx')


# 

# In[123]:


"""# Fréquence Reddit"""
pd.read_csv("data/reddit_posts_avec_mot.csv").groupby("annee")["annee"].count().reset_index(name="nombre")


# In[124]:


df_freq_annee = pd.read_csv("data/reddit_posts_avec_mot.csv").groupby("annee")["annee"].count().reset_index(name="nombre")
df_freq_annee


# In[125]:


from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

# Charger le fichier Excel
wb = load_workbook("data/tdb_obs.xlsx")

# Accéder à la feuille "Feuil1"
ws_feuil1 = wb["Feuil1"]

# Créer ou accéder à la feuille "frequence_reddit"
if "frequence_reddit" in wb.sheetnames:
    ws_reddit = wb["frequence_reddit"]
else:
    ws_reddit = wb.create_sheet(title="frequence_reddit")

# Position pour coller les données dans "frequence_reddit"
start_col = 2  # B
start_row = 1  # Ligne 1

# Écrire les en-têtes dans "frequence_reddit"
ws_reddit.cell(row=start_row, column=start_col, value="Année")
ws_reddit.cell(row=start_row, column=start_col + 1, value="Nombre")

# Coller les données dans "frequence_reddit"
for idx, row in enumerate(df_freq_annee.itertuples(index=False), start=start_row + 1):
    ws_reddit.cell(row=idx, column=start_col, value=row.annee)
    ws_reddit.cell(row=idx, column=start_col + 1, value=row.nombre)

# Créer le LineChart dans "Feuil1" en utilisant les données de "frequence_reddit"
chart = LineChart()
data = Reference(ws_reddit, min_col=start_col + 1, min_row=start_row, max_row=start_row + len(df_freq_annee))
labels = Reference(ws_reddit, min_col=start_col, min_row=start_row + 1, max_row=start_row + len(df_freq_annee))

chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)
chart.title = "Évolution du nombre de posts par année"
chart.y_axis.title = "Nombre de posts"
chart.x_axis.title = "Année"

# Position du graphique dans "Feuil1"
chart_position = 'H7'
ws_feuil1.add_chart(chart, chart_position)

# Sauvegarde
wb.save("data/tdb_obs.xlsx")


# # Score de popularité

# In[ ]:





# In[126]:


"""score de popularité"""
pd.read_csv('data/score_de_popularite.csv')[['score_de_popularite_twitter', 'score_de_popularite_reddit']]


# In[127]:


from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def ecrire_scores_popularite(ws, df, start_cell):
        # Couleurs
        rouge = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rouge clair
        vert = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Vert clair

        # Coordonnées de départ
        col_letter = start_cell[0]
        row_index = int(start_cell[1:])

        # Écrire le titre
        ws[f"{col_letter}{row_index}"] = "Score de popularité"
        ws.merge_cells(f"{col_letter}{row_index}:{chr(ord(col_letter)+1)}{row_index}")

        # En-têtes
        ws[f"{col_letter}{row_index + 1}"] = "Twitter"
        ws[f"{chr(ord(col_letter)+1)}{row_index + 1}"] = "Reddit"

        # Remplissage des données avec couleurs
        for i, row in df.iterrows():
            twitter_val = row["score_de_popularite_twitter"]
            reddit_val = row["score_de_popularite_reddit"]

            row_excel = row_index + 2 + i
            cell_twitter = f"{col_letter}{row_excel}"
            cell_reddit = f"{chr(ord(col_letter)+1)}{row_excel}"

            ws[cell_twitter] = twitter_val
            ws[cell_reddit] = reddit_val

            # Remplissage conditionnel Twitter
            if twitter_val >= 50:
                ws[cell_twitter].fill = vert
            else:
                ws[cell_twitter].fill = rouge

            # Remplissage conditionnel Reddit
            if reddit_val >= 50:
                ws[cell_reddit].fill = vert
            else:
                ws[cell_reddit].fill = rouge

ecrire_scores_popularite(
    wb["Feuil1"],
    pd.read_csv('data/score_de_popularite.csv')[['score_de_popularite_twitter', 'score_de_popularite_reddit']],
    "J2"
)

wb.save("data/tdb_obs.xlsx")


# # Articles

# In[128]:


"""articles"""
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference

def ajouter_barplot_par_journal(df, wb, cellule_depart="M1"):
    # 1. Créer ou récupérer la feuille "articles"
    if "articles" in wb.sheetnames:
        ws_data = wb["articles"]
    else:
        ws_data = wb.create_sheet("articles")

    # 2. Groupby journal
    counts = df["journal"].value_counts().reset_index()
    counts.columns = ["journal", "nombre"]

    # 3. Écrire les données dans "articles"
    ws_data.cell(row=1, column=1, value="Journal")
    ws_data.cell(row=1, column=2, value="Nombre")

    for i, row in counts.iterrows():
        ws_data.cell(row=2 + i, column=1, value=row["journal"])
        ws_data.cell(row=2 + i, column=2, value=row["nombre"])

    # 4. Créer le graphique dans Feuil1 (ou "Analyse")
    ws_chart = wb["Feuil1"]

    chart = BarChart()
    chart.title = "Nombre d'articles par journal"
    chart.x_axis.title = "Journal"
    chart.y_axis.title = "Nombre"

    data = Reference(ws_data, min_col=2, min_row=1, max_row=1 + len(counts))
    categories = Reference(ws_data, min_col=1, min_row=2, max_row=1 + len(counts))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    # 5. Ajouter le graphique à la feuille cible
    ws_chart.add_chart(chart, cellule_depart)

df = pd.read_csv("data/actualite_avec_mot.csv")

ajouter_barplot_par_journal(df, wb, cellule_depart="M1")

wb.save("data/tdb_obs.xlsx")


# # Tableau

# In[129]:


"""génération des tableaux de tdb"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter, column_index_from_string

def export_text_table_to_excel(df, filename="data/tdb_obs.xlsx", sheet_title="Feuil1", table_title="Figure", start_cell='A1'):
    ws = wb['Feuil1']

    # Récupérer coordonnées de départ
    start_col_letter = ''.join(filter(str.isalpha, start_cell))
    start_row = int(''.join(filter(str.isdigit, start_cell)))

    start_col = column_index_from_string(start_col_letter)

    # Bordure noire
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Si titre, on l’insère juste au-dessus
    if table_title:
        ws.merge_cells(
            start_row=start_row, start_column=start_col, 
            end_row=start_row, end_column=start_col + len(df.columns) - 1
        )
        cell = ws.cell(row=start_row, column=start_col, value=table_title)
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        start_row += 1  # Décaler les en-têtes

    # En-têtes
    for col_idx, col_name in enumerate(df.columns):
        cell = ws.cell(row=start_row, column=start_col + col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Données
    for row_idx, (_, row) in enumerate(df.iterrows()):
        for col_idx, value in enumerate(row):
            cell = ws.cell(row=start_row + 1 + row_idx, column=start_col + col_idx, value=str(value))
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Ajuster largeur des colonnes
    for col_idx, col_name in enumerate(df.columns):
        max_len = max([len(str(col_name))] + [len(str(val)) for val in df.iloc[:, col_idx]])
        col_letter = get_column_letter(start_col + col_idx)
        ws.column_dimensions[col_letter].width = min(50, max_len + 2)

    # Ajuster hauteur des lignes
    for row in range(start_row + 1, start_row + 1 + len(df)):
        ws.row_dimensions[row].height = 60

    wb.save(filename)


# In[130]:


# Charger les fichiers CSV
df_tweets = pd.read_csv("data/tweets_avec_mot.csv")
df_reddit = pd.read_csv("data/reddit_posts_avec_mot.csv")
df_medias = pd.read_csv("data/actualite_avec_mot.csv")

import random

def choix_aleatoire(df, colonne):
    if df.empty or colonne not in df.columns or df[colonne].dropna().empty:
        return "Aucun résultat"
    else:
        return random.choice(df[colonne].dropna().values)

df_verbatim = pd.DataFrame.from_dict({
    "Twitter": [choix_aleatoire(df_tweets, "tweets")],
    "Reddit": [choix_aleatoire(df_reddit, "selftext")],
    "Médias": [choix_aleatoire(df_medias, "description")]
})


df_verbatim


# In[131]:


export_text_table_to_excel(df_verbatim, sheet_title="Feuil1", table_title="Verbatim", start_cell='G19')


# In[132]:


df_definitions = pd.read_csv('data/definitions.csv').head(5)
export_text_table_to_excel(df_definitions, sheet_title="Feuil1", table_title="Définitions", start_cell='A1')


# In[133]:


df_citations = pd.read_csv("data/citations.csv").head(5)
df_citations


# In[134]:


export_text_table_to_excel(df_citations, sheet_title="Feuil1", table_title="Citations", start_cell='C1')


# In[135]:


export_text_table_to_excel(
    pd.read_csv("data/prononciation.csv"),
    sheet_title="Feuil1",
    table_title="Prononciation",
    start_cell='E1')

export_text_table_to_excel(
    pd.read_csv("data/etymologies.csv"),
    sheet_title="Feuil1",
    table_title="Prononciation",
    start_cell='E4')

